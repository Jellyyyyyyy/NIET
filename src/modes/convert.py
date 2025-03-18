import json
import re
import csv
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# Module-level constant for IP matching.
IP_REGEX = re.compile(r"^\d{1,3}(?:\.\d{1,3}){3}$")


def remove_duplicates(rows, keys):
    """
    Remove duplicate rows from a list of dictionaries.
    Duplicates are defined as rows that have identical values for every key in 'keys'.
    """
    unique_rows = []
    seen = set()
    for row in rows:
        key_tuple = tuple(row.get(key, "").strip() for key in keys)
        if key_tuple not in seen:
            seen.add(key_tuple)
            unique_rows.append(row)
    return unique_rows


def create_table_and_style(sheet, table_name, config=None):
    """
    Create an Excel table on the given sheet and apply header fill, font, and border styles.
    If the sheet has only a header row (no data), a dummy row is added so the table range is valid.
    """
    if config is None:
        config = {}
    
    table_style = config.get("table_style", "TableStyleMedium9")
    show_first_column = config.get("show_first_column", False)
    show_last_column = config.get("show_last_column", False)
    show_row_stripes = config.get("show_row_stripes", False)
    show_column_stripes = config.get("show_column_stripes", False)
    header_fill = config.get("header_fill", "A5A5A5")
    header_fill_type = config.get("header_fill_type", "solid")
    header_font_color = config.get("header_font_color", "000000")
    cell_fill = config.get("cell_fill", "ffffff")
    cell_fill_type = config.get("cell_fill_type", "solid")
    cell_border = config.get("cell_border", "thin")
    cell_font_color = config.get("cell_font_color", "000000")

    if sheet.max_row < 2:
        # Insert a dummy row with empty strings for each column
        sheet.append([""] * sheet.max_column)
    
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    table_range = f"A1:{get_column_letter(max_col)}{max_row}"
    tab = Table(displayName=table_name.replace(" ", ""), ref=table_range)
    style = TableStyleInfo(name=table_style, showFirstColumn=show_first_column,
                           showLastColumn=show_last_column, showRowStripes=show_row_stripes, showColumnStripes=show_column_stripes)
    tab.tableStyleInfo = style
    sheet.add_table(tab)
    
    thin_border = Border(left=Side(style=cell_border), right=Side(style=cell_border),
                        top=Side(style=cell_border), bottom=Side(style=cell_border))
    
    header_fill = PatternFill(start_color=header_fill, end_color=header_fill, fill_type=header_fill_type)
    for cell in sheet[1]:
        cell.border = thin_border
        cell.fill = header_fill
        cell.font = Font(color=header_font_color)
        
    cell_fill = PatternFill(start_color=cell_fill, end_color=cell_fill, fill_type=cell_fill_type)
    for row in sheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            cell.fill = cell_fill
            cell.font = Font(color=cell_font_color)
            

def write_csv_sheet(sheet, dataset, csv_header, sheet_options, extract_config=None):
    global COUNT
    if extract_config is None:
        extract_config = {}
    
    headers_to_remove = [h.lower() for h in sheet_options.get("headers_to_remove", [])]
    csv_header_no = [col for col in csv_header if col.lower() not in headers_to_remove]
    
    # Build the new header.
    new_header = []
    for col in csv_header_no:
        new_header.append(col)
        if col.lower() == sheet_options.get("insert_mapping_columns_after", "Risk").lower():
            if sheet_options.get("mapping_columns", None) is not None:
                new_header.extend(sheet_options.get("mapping_columns", []))
    
    if extract_config.get("extract_column_name", None) is not None:
        new_header.extend([extract_config.get("extract_column_name")])
    
    additional_columns = sheet_options.get("additional_columns", None)
    if additional_columns is not None:
        new_header.extend(additional_columns.keys())
    
    sheet.append(new_header)
    
    # Process each row in the dataset.
    for row in dataset:
        base_row = []
        for col in csv_header_no:
            base_row.append(row.get(col, ""))
            if col.lower() == sheet_options.get("insert_mapping_columns_after", "Risk").lower():
                host_val = row.get(sheet_options.get("host_column_name", "Host"), "").strip()
                
                if IP_REGEX.match(host_val):
                    mapping_val_row = sheet_options.get("ip_match", [0, 0, 1, 0, 0, 0])[:]
                else:
                    mapping_val_row = sheet_options.get("not_ip_match", [1, 0, 0, 0, 0, 0])[:]
                    
                for i, val in enumerate(mapping_val_row):
                    if val == 1:
                        mapping_val_row[i] = host_val
                base_row.extend(mapping_val_row)
                
        # If extra processing is requested, create one row per extra value.
        if extract_config.get("extract_column_name", None) is not None:
            extra_values = extract_information(row, extract_config)
            for extra in extra_values:
                extra_row = base_row[:]
                extra_row.extend([extra, *[value for value in additional_columns.values()]])
                sheet.append(extra_row)
        else:
            base_row.extend([value for value in additional_columns.values()])
            sheet.append(base_row)


def add_lookup_formulas(sheet, hosts_sheet, sheet_options):
    """
    Insert lookup formulas in the host-lookup columns of the given sheet using data from the Hosts sheet.
    """
    # Extract configuration from sheet_options
    mapping_cols = sheet_options.get("mapping_columns", [])
    mapping_formulas = sheet_options.get("mapping_columns_formula", {})
    
    # Build a dictionary mapping each configured column name to its column index in the sheet
    headers = [cell.value for cell in sheet[1]]
    col_indices = {}
    for col in mapping_cols:
        try:
            col_indices[col] = headers.index(col) + 1
        except ValueError:
            # Skip if the expected column is not found
            continue

    # Determine the number of rows in the Hosts sheet (to define lookup ranges)
    host_rows = hosts_sheet.max_row
    if host_rows < 2:
        return

    # Process each data row in the sheet
    for i in range(2, sheet.max_row + 1):
        # Build a dynamic replacements dictionary for the current row.
        # For each mapping column, we create keys like {ColumnName_cell} and {columnname_cell}
        # to handle different naming conventions in the formula templates.
        replacements = {"host_rows": str(host_rows), "hosts_table_name": sheet_options.get("hosts_table_name", "Hosts")}
        for col in mapping_cols:
            if col in col_indices:
                cell_ref = f"{get_column_letter(col_indices[col])}{i}"
                replacements[f"{col}_cell"] = cell_ref
                replacements[f"{col.lower()}_cell"] = cell_ref  # Cover lower-case placeholder

        # Loop over each formula in the config and apply substitutions.
        # The key in mapping_formulas is the target column where the formula should be placed.
        for target_col, formula_template in mapping_formulas.items():
            if target_col not in col_indices:
                continue  # Skip if the target column isn't in the sheet
            formula = formula_template
            # Replace every placeholder in the formula with the dynamic cell references
            for placeholder, value in replacements.items():
                formula = formula.replace(f"{{{placeholder}}}", value)
            # Assign the computed formula to the appropriate cell in the target column
            sheet.cell(row=i, column=col_indices[target_col]).value = formula


def add_status_validation(sheet, config=None):
    """
    Add data validation for the "Status" column on the given sheet.
    Acceptable values: "-", "Open", "On-going", "Closed".
    """
    if config is None:
        config = {}
        
    DEFAULT_STATUS_MAP = {
        "-": "ffffff",
        "Open": "FF0000",
        "On-going": "FFC7CE",
        "Closed": "FFEB9C",
        "Declared": "24FC03"
    }
        
    status_map = config.get("status_map", DEFAULT_STATUS_MAP)
    
    headers = [cell.value for cell in sheet[1]]
    try:
        status_idx = headers.index("Status") + 1
    except ValueError:
        return
    max_row = sheet.max_row
    # Only apply data validation if there are data rows.
    if max_row < 2:
        return
    dv = DataValidation(type="list", formula1=f'"{",".join(status_map.keys())}"', allow_blank=False)
    dv.error = 'Select a value from the list'
    dv.errorTitle = 'Invalid Entry'
    dv_range = f"{get_column_letter(status_idx)}2:{get_column_letter(status_idx)}{max_row}"
    sheet.add_data_validation(dv)
    dv.add(dv_range)


def add_conditional_formatting(sheet, config=None):
    """
    Add conditional formatting to the "Status" column.
    Applies a red fill for "Open" and an orange fill for "On-going".
    """
    if config is None:
        config = {}
        
    DEFAULT_STATUS_MAP = {
        "-": "ffffff",
        "Open": "FF0000",
        "On-going": "FFC7CE",
        "Closed": "FFEB9C",
        "Declared": "24FC03"
    }
        
    status_map = config.get("status_map", DEFAULT_STATUS_MAP)
    status_fill_type = config.get("status_fill_type", "solid")
    
    headers = [cell.value for cell in sheet[1]]
    try:
        status_idx = headers.index("Status") + 1
    except ValueError:
        return
    max_row = sheet.max_row
    if max_row < 2:
        return
    data_range = f"A2:{get_column_letter(sheet.max_column)}{max_row}"
    
    for status, color in status_map.items():
        formula = f'=${get_column_letter(status_idx)}2="{status}"'
        fill = PatternFill(start_color=color, end_color=color, fill_type=status_fill_type)
        sheet.conditional_formatting.add(data_range, FormulaRule(formula=[formula], fill=fill))


def extract_information(row, config):
    regex_flags = re.IGNORECASE if config.get("case_insensitive") else 0

    # Process lookup_columns to extract lookup values
    lookup_col_patterns = config.get("lookup_columns", [])
    compiled_lookup_col_patterns = [re.compile(pat, regex_flags) for pat in lookup_col_patterns]

    lookup_values = []
    for col_name, value in row.items():
        for pattern in compiled_lookup_col_patterns:
            if pattern.search(col_name):
                lookup_values.append(value.strip())
                break

    # Determine OS type based on lookup values
    for os_type in ["linux", "windows"]:
        lookup_patterns = config.get(os_type, {}).get("lookup_values", {})
        compiled_lookup_patterns = [re.compile(pat, regex_flags) for pat in lookup_patterns]

        if any(any(p.search(val) for p in compiled_lookup_patterns) for val in lookup_values):
            user_type = os_type
            break
    else:
        return []
    


    # Extract relevant text from specified columns
    extract_col_patterns = config.get("extract_columns", [])
    compiled_extract_col_patterns = [re.compile(pat, regex_flags) for pat in extract_col_patterns]

    extracted_text_parts = []
    for col_name, value in row.items():
        for pattern in compiled_extract_col_patterns:
            if pattern.search(col_name):
                extracted_text_parts.append(value.strip())
                break
    extracted_text = "\n".join(extracted_text_parts)

    # Process OS-specific extraction patterns
    extraction_conf = config.get(user_type, {}).get("extraction", {})
    user_regexes = extraction_conf.get("regex", [])
    compiled_user_regexes = [re.compile(pat, regex_flags) for pat in user_regexes]

    exclude_patterns = extraction_conf.get("exclude", [])
    compiled_exclude_patterns = [re.compile(pat, regex_flags) for pat in exclude_patterns]

    results = []
    for line in extracted_text.splitlines():
        line = line.strip()
        if not line or any(p.search(line) for p in compiled_exclude_patterns):
            continue
        for regex in compiled_user_regexes:
            match = regex.match(line)
            if match:
                extracted_value = match.group(1).strip() if match.groups() else line
                results.append(extracted_value)
                break

    return results


def hide_and_autowidth_columns(sheet, allowed_columns, auto_columns):
    """
    Hide all columns not listed in allowed_columns.
    For columns in auto_columns, adjust the width based on the maximum cell length.
    """
    header = [cell.value for cell in sheet[1]]
    for idx, col_name in enumerate(header, start=1):
        col_letter = get_column_letter(idx)
        if col_name not in allowed_columns:
            sheet.column_dimensions[col_letter].hidden = True
        elif col_name in auto_columns:
            max_length = 0
            for cell in sheet[col_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max_length + 2


def add_risk_font_formatting(sheet, sheet_options):
    """
    Add conditional formatting rules to change the font color based on the risk level.
    """
    headers = [cell.value for cell in sheet[1]]
    try:
        idx = headers.index(sheet_options.get("text_format_column", "Risk")) + 1
    except ValueError:
        return
    max_row = sheet.max_row
    if max_row < 2:
        return
    col_letter = get_column_letter(idx)
    data_range = f"{col_letter}2:{col_letter}{max_row}"
    for val, color in sheet_options.get("text_format", {}).items():
        rule = FormulaRule(formula=[f'=${col_letter}2="{val}"'], font=Font(color=color))
        sheet.conditional_formatting.add(data_range, rule)


def nessus_convert(csv_filename: str, excel_filename: str, logger=None, software_exclusion_keywords=None, flags=None, config_path=None):
    """
    Convert the given CSV file into an Excel workbook.
    Six sheets are created (in order):
      1. "Hosts": Unique hosts from the CSV with columns "Hostname", "IP", "OS".
      2. "Vulnerabilities": Processed CSV rows with Risk in {"None","Low","Medium","High","Critical"}.
         After removing "Host" and "Plugin Output", immediately after "Risk" insert six new host-lookup columns:
         "Hostname_original", "Hostname", "IP_original", "IP", "OS_original", "OS", then append "Status" (set to "Open")
         and "Remarks".
      3. "Compliance": Processed similarly for rows with Risk in {"FAILED","WARNING"}.
      4. "Open Ports": Processed similarly for rows where Name equals "Netstat Portscanner (SSH)".
      5. "Users": For rows where Name is "Linux User List Enumeration" or "Enumerate Users via WMI",
         extract individual users from "Plugin Output" and for each create a new row. The header is built as above,
         then extra columns "User", "Status", "Remarks" are appended. Additionally, the six host-lookup columns
         are included and filled with structured formulas.
      6. "Installed Software": For rows where Name contains "Software Enumeration" (case insensitive)
         and does not contain any exclusion keyword (default: ["startup", "start-up", "os"]),
         for each host (only the first matching row per host is kept) the "Plugin Output" is split (by newlines)
         and each extracted line is written as a new row. The header is built as above but with extra columns
         "Installed Programs", "Status", "Remarks". Again, the six host-lookup columns are included and filled.
    
    All CSV-derived sheets are converted into Excel tables with header cells styled with a solid fill (RGB A5A5A5)
    with black text and full thin borders. Data validation is added so that the "Status" column only accepts
    "-", "Open", "On-going", or "Closed". In Vulnerabilities, Compliance and Open Ports the host-lookup columns
    are filled with lookup formulas (using A1-style references to the Hosts sheet). For Users and Installed Software,
    the host-lookup columns are present and then overwritten with structured reference formulas.
    
    Finally, the visible columns are limited (the rest are hidden) as follows:
         - Hosts: ["Hostname", "IP", "OS"]
         - Vulnerabilities: ["Risk", "Hostname", "IP", "OS", "Name", "Synopsis", "Description", "Solution", "See Also", "Status", "Remarks"]
         - Compliance: ["Risk", "Hostname", "IP", "OS", "Name", "Description", "Solution", "See Also", "Status", "Remarks"]
         - Open Ports: ["Hostname", "IP", "OS", "Protocol", "Port", "Status", "Remarks"]
         - Users: ["Hostname", "IP", "OS", "User", "Status", "Remarks"]
         - Installed Software: ["Hostname", "IP", "OS", "Installed Programs", "Status", "Remarks"]
    """
    if config_path:
        with open(config_path, "r") as f:
            config = json.load(f).get("excel_config", {})
            config = {**generate_default_config(), **config}
    else:
        config = generate_default_config()
    
    if software_exclusion_keywords is not None:
        config.setdefault("sheets", {}).setdefault("installed_software", {}).setdefault("filter_exclude", []).extend(software_exclusion_keywords)

    hosts_set = set()
    try:
        with open(csv_filename, newline="", encoding="utf-8") as csvfile:
            reader = csv.DictReader(csvfile)
            csv_header = reader.fieldnames
            rows = list(reader)
            
            if config.get("remove_duplicates", True):
                rows = remove_duplicates(rows, csv_header)
                
            sheets = {sheet_name.get("sheet_name"): [] for sheet_name in config.get("sheets", {}).values()}
            for row in rows:
                host_val = row.get(config.get("host_column_name", "Host"), "").strip()
                if host_val:
                    hosts_set.add(host_val)
                
                for sheet_config in config.get("sheets", {}).values():
                    sheet_name = sheet_config.get("sheet_name")
                    regex_flags = re.IGNORECASE if sheet_config.get("case_insensitive", False) else 0

                    # Compile the patterns
                    compiled_col_filter_patterns = [re.compile(pat, regex_flags) for pat in sheet_config.get("column_filter_lookup", [])]
                    compiled_filter_patterns = [re.compile(pat, regex_flags) for pat in sheet_config.get("filter", [])]
                    compiled_filter_exclude_patterns = [re.compile(pat, regex_flags) for pat in sheet_config.get("filter_exclude", [])]

                    # Helper function: returns True if any pattern in the list matches the text
                    def matches_any(patterns, text):
                        return any(pattern.search(text) for pattern in patterns)

                    if any(
                        matches_any(compiled_filter_patterns, value) and not matches_any(compiled_filter_exclude_patterns, value)
                        for col_name, value in row.items() if matches_any(compiled_col_filter_patterns, col_name)
                    ):
                        sheets[sheet_name].append(row)
        
    except Exception as e:
        if logger:
            logger.error(f"Error reading CSV file {csv_filename}: {e}")
        return

    # --- Step 2. Create workbook and Hosts sheet ---
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    hosts_sheet = wb.create_sheet(config.get("hosts", {}).get("sheet_name", "Hosts"))
    hosts_header = config.get("hosts", {}).get("headers", ["Hostname", "IP", "OS"])
    hosts_sheet.append(hosts_header)
    for host in sorted(hosts_set):
        if IP_REGEX.match(host):
            row = config.get("hosts", {}).get("ip_match", ["", 1, ""])[:]
        else:
            row = config.get("hosts", {}).get("not_ip_match", [1, "", ""])[:]
            
        for i, val in enumerate(row):
            if val == 1:
                row[i] = host

        hosts_sheet.append(row)

    create_table_and_style(hosts_sheet, "Hosts", config.get("table", {}))
    hide_and_autowidth_columns(hosts_sheet, config.get("hosts", {}).get("visible_columns", []), [])

    for sheet_config in config.get("sheets", {}).values():
        sheet_name = sheet_config.get("sheet_name")
        sheet = wb.create_sheet(sheet_name)
            
        write_csv_sheet(sheet, sheets[sheet_name], csv_header, config.get("sheet_options", {}), sheet_config.get("extract_config", {}))

        create_table_and_style(sheet, sheet_name, config.get("table", {}))
        add_lookup_formulas(sheet, hosts_sheet, config.get("sheet_options", {}))
        add_status_validation(sheet, config.get("status", {}))
        add_conditional_formatting(sheet, config.get("status", {}))
        hide_and_autowidth_columns(sheet, sheet_config.get("visible_columns", []), sheet_config.get("auto_width_columns", []))
        add_risk_font_formatting(sheet, sheet_config)

    try:
        wb.save(excel_filename)
        if logger:
            logger.info(f"Excel file created successfully at: {excel_filename}")
    except Exception as e:
        if logger:
            logger.error(f"Error saving Excel file {excel_filename}: {e}")


def generate_default_config():
    return  {
    "remove_duplicates": True,
    "table": {
      "table_style": "TableStyleMedium9",
      "show_first_column": False,
      "show_last_column": False,
      "show_row_stripes": False,
      "show_column_stripes": False,
      "header_fill": "A5A5A5",
      "header_fill_type": "solid",
      "header_font_color": "000000",
      "cell_fill": "ffffff",
      "cell_fill_type": "solid",
      "cell_border": "thin",
      "cell_font_color": "000000"
    },
    "status": {
      "status_map": {
        "-": "ffffff",
        "Open": "FFC7CE",
        "On-going": "FFEB9C",
        "Closed": "FFFFFF",
        "Declared": "82F073"
      },
      "status_fill_type": "solid"
    },
    "hosts": {
      "sheet_name": "Hosts",
      "table_name": "Hosts",
      "headers": ["Hostname", "IP", "OS"],
      "visible_columns": ["Hostname", "IP", "OS"],
      "ip_match": ["", 1, ""],
      "not_ip_match": [1, "", ""]
    },
    "sheet_options": {
      "host_column_name": "Host",
      "insert_mapping_columns_after": "Risk",
      "mapping_columns": ["Hostname_original", "Hostname", "IP_original", "IP", "OS_original", "OS"],
      "ip_match": ["", "", 1, "", "", ""],
      "not_ip_match": [1, "", "", "", "", ""],
      "mapping_columns_formula": {
        "Hostname": "=IF(ISBLANK({Hostname_original_cell}), IF(ISBLANK(INDEX({hosts_table_name}!$A$2:$A${host_rows}, MATCH({IP_original_cell}, {hosts_table_name}!$B$2:$B${host_rows}, 0))), \"\", INDEX({hosts_table_name}!$A$2:$A${host_rows}, MATCH({IP_original_cell}, {hosts_table_name}!$B$2:$B${host_rows}, 0))), {Hostname_original_cell})",
        "IP": "=IF(ISBLANK({IP_original_cell}), IF(ISBLANK(INDEX({hosts_table_name}!$B$2:$B${host_rows}, MATCH({Hostname_original_cell}, {hosts_table_name}!$A$2:$A${host_rows}, 0))), \"\", INDEX({hosts_table_name}!$B$2:$B${host_rows}, MATCH({Hostname_original_cell}, {hosts_table_name}!$A$2:$A${host_rows}, 0))), {IP_original_cell})",
        "OS": "=IF(ISBLANK({OS_original_cell}), IF(ISBLANK(INDEX({hosts_table_name}!$C$2:$C${host_rows}, MATCH({ip_cell}, {hosts_table_name}!$B$2:$B${host_rows}, 0))), \"\", INDEX({hosts_table_name}!$C$2:$C${host_rows}, MATCH({ip_cell}, {hosts_table_name}!$B$2:$B${host_rows}, 0))), {OS_original_cell})"
      },
      "additional_columns": {"Status": "Open", "Remarks": ""},
      "headers_to_remove": ["Host"]
    },
    "sheets": {
      "vulnerabilities":{
        "sheet_name": "Vulnerabilities",
        "case_insensitive": True,
        "filter": ["^None$", "^Low$", "^Medium$", "^High$", "^Critical$"],
        "column_filter_lookup": ["^Risk$"],
        "filter_exclude": [],
        "visible_columns": ["Risk", "Hostname", "IP", "OS", "Name", "Synopsis", "Description", "Solution", "See Also", "Plugin Output", "Status", "Remarks"],
        "auto_width_columns": [],
        "text_format_column": "Risk",
        "text_format": {
          "Low": "0000FF",
          "Medium": "A0522D",
          "High": "FF0000",
          "Critical": "8B0000"
        }
      },
      "compliance":{
        "sheet_name": "Compliance",
        "case_insensitive": True,
        "filter": ["^WARNING$", "^FAILED$"],
        "column_filter_lookup": ["^Risk$"],
        "filter_exclude": [],
        "visible_columns": ["Risk", "Hostname", "IP", "OS", "Name", "Synopsis", "Description", "Solution", "See Also", "Plugin Output", "Status", "Remarks"],
        "auto_width_columns": [],
        "text_format_column": "Risk",
        "text_format": {
          "WARNING": "A0522D",
          "FAILED": "FF0000"
        }
      },
      "open_ports":{
        "sheet_name": "Open Ports",
        "case_insensitive": True,
        "filter": ["Netstat Portscanner \(SSH\)", "Netstat Portscanner \(WMI\)"],
        "column_filter_lookup": ["^Name$"],
        "filter_exclude": [],
        "visible_columns": ["Hostname", "IP", "OS", "Protocol", "Port", "Status", "Remarks"],
        "auto_width_columns": []
      },
      "users":{
        "sheet_name": "Users",
        "case_insensitive": True,
        "filter": ["^Linux User List Enumeration$", "^Enumerate Users via WMI$"],
        "column_filter_lookup": ["^Name$"],
        "filter_exclude": [],
        "visible_columns": ["Hostname", "IP", "OS", "User", "Status", "Remarks"],
        "auto_width_columns": [],
        "extract_config": {
          "case_insensitive": True,
          "lookup_columns": ["^Name$"],
          "extract_columns": ["^Plugin Output$"],
          "extract_column_name": "User",
          "linux": {
              "lookup_values": ["^Linux User List Enumeration$"],
              "extraction": {
                  "regex": ["^User\s*:\s*(.+)$"],
                  "exclude": []
              }
          },
          "windows": {
              "lookup_values": ["^Enumerate Users via WMI$"],
              "extraction": {
                  "regex": ["^Name\s*:\s*(.+)$"],
                  "exclude": ["no\.?\s*of\s*users"]
              }
          }
        }

      },
      "installed_software":{
        "sheet_name": "Installed Software",
        "case_insensitive": True,
        "filter": ["software enumeration"],
        "column_filter_lookup": ["^Name$"],
        "filter_exclude": ["identification", "startup", "start-up"],
        "visible_columns": ["Hostname", "IP", "OS", "Installed Program", "Status", "Remarks"],
        "auto_width_columns": [],

        "extract_config": {
          "case_insensitive": True,
          "lookup_columns": ["^Name$"],
          "extract_columns": ["^Plugin Output$"],
          "extract_column_name": "Installed Program",
          "linux": {
              "lookup_values": ["ssh"],
              "extraction": {
                  "regex": [".*"],
                  "exclude": ["list of packages", "^-+"]
              }
          },
          "windows": {
              "lookup_values": ["microsoft windows installed software enumeration"],
              "extraction": {
                  "regex": [".*installed on.*"],
                  "exclude": ["the following software"]
              }
          }
        }
      }
    }
  }