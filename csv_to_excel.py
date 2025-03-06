import re
import csv
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# Module-level constant for IP matching.
IP_REGEX = re.compile(r"^\d{1,3}(?:\.\d{1,3}){3}$")


def remove_duplicates(rows, keys):
    """
    Remove duplicate rows from a list of dictionaries.
    Duplicates are defined as rows that have identical values for every key in 'keys'.
    """
    unique_rows = []
    seen = set()
    for row in rows:
        key_tuple = tuple(row.get(key, "").strip() for key in keys)
        if key_tuple not in seen:
            seen.add(key_tuple)
            unique_rows.append(row)
    return unique_rows


def create_table_and_style(sheet, table_name):
    """
    Create an Excel table on the given sheet and apply header fill, font, and border styles.
    """
    max_row = sheet.max_row
    max_col = sheet.max_column
    table_range = f"A1:{get_column_letter(max_col)}{max_row}"
    tab = Table(displayName=table_name.replace(" ", ""), ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    tab.tableStyleInfo = style
    sheet.add_table(tab)
    header_fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="000000")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border


def write_csv_sheet(sheet, dataset, csv_header, extra_col_name=None, extract_func=None):
    """
    Write rows to a CSV sheet with optional extra processing.
    
    The function builds the CSV header by:
      - Removing the "Host" and "Plugin Output" columns from csv_header.
      - Inserting six host-lookup columns ("Hostname_original", "Hostname", "IP_original", "IP", "OS_original", "OS")
        immediately after the "Risk" column.
      - Appending extra_col_name (if provided) followed by "Status" and "Remarks", or just "Status" and "Remarks" otherwise.
    
    For each row in dataset:
      - A base row is built using the filtered header columns.
      - When processing the "Risk" column, the function uses the value from "Host" (after stripping whitespace) 
        to determine host-lookup values. If the host value matches IP_REGEX (assumed to be defined globally), it extends
        the base row with ["", "", host_val, "", "", ""]; otherwise, it extends it with [host_val, "", "", "", "", ""].
      - If extract_func is provided (and hence extra_col_name is also provided), the function calls extract_func(row)
        to obtain extra values and then appends one row per extra value (i.e. base_row + [extra, "Open", ""]).
      - Otherwise, it simply appends a single row per dataset row with base_row extended by ["Open", ""].
    
    Parameters:
      sheet: The sheet object (e.g., a list or worksheet) where rows will be appended.
      dataset: An iterable of dictionaries, each representing a row of data.
      csv_header: A list of column names for the CSV.
      extra_col_name (optional): The header name for the extra column. Must be provided with extract_func.
      extract_func (optional): A function that takes a row and returns an iterable of extra values.
    
    Raises:
      ValueError: If only one of extra_col_name or extract_func is provided.
    """
    # Check that both or neither of the extra parameters are provided.
    if (extra_col_name is None) != (extract_func is None):
        raise ValueError("Both extra_col_name and extract_func must be provided together")
    
    # Remove "Host" and "Plugin Output" columns from the header.
    csv_header_no = [col for col in csv_header if col.lower() not in {"host", "plugin output"}]
    
    # Build the new header.
    new_header = []
    for col in csv_header_no:
        new_header.append(col)
        if col.lower() == "risk":
            new_header.extend(["Hostname_original", "Hostname", "IP_original", "IP", "OS_original", "OS"])
    
    # Append extra_col_name if provided, then "Status" and "Remarks".
    if extra_col_name is not None:
        new_header.extend([extra_col_name, "Status", "Remarks"])
    else:
        new_header.extend(["Status", "Remarks"])
    
    sheet.append(new_header)
    
    # Process each row in the dataset.
    for row in dataset:
        base_row = []
        for col in csv_header_no:
            base_row.append(row.get(col, ""))
            if col.lower() == "risk":
                host_val = row.get("Host", "").strip()
                # Assumes IP_REGEX is defined in the global scope.
                if IP_REGEX.match(host_val):
                    base_row.extend(["", "", host_val, "", "", ""])
                else:
                    base_row.extend([host_val, "", "", "", "", ""])
        # If extra processing is requested, create one row per extra value.
        if extract_func is not None:
            extra_values = extract_func(row)
            for extra in extra_values:
                sheet.append(base_row + [extra, "Open", ""])
        else:
            base_row.extend(["Open", ""])
            sheet.append(base_row)


def add_lookup_formulas(sheet, hosts_sheet):
    """
    Insert lookup formulas in the host-lookup columns of the given sheet using data from the Hosts sheet.
    """
    headers = [cell.value for cell in sheet[1]]
    try:
        h_orig_idx = headers.index("Hostname_original") + 1
        h_idx = headers.index("Hostname") + 1
        ip_orig_idx = headers.index("IP_original") + 1
        ip_idx = headers.index("IP") + 1
        os_orig_idx = headers.index("OS_original") + 1
        os_idx = headers.index("OS") + 1
    except ValueError:
        return
    host_rows = hosts_sheet.max_row
    for i in range(2, sheet.max_row + 1):
        h_orig_cell = f"{get_column_letter(h_orig_idx)}{i}"
        h_cell = f"{get_column_letter(h_idx)}{i}"
        ip_orig_cell = f"{get_column_letter(ip_orig_idx)}{i}"
        ip_cell = f"{get_column_letter(ip_idx)}{i}"
        os_orig_cell = f"{get_column_letter(os_orig_idx)}{i}"
        os_cell = f"{get_column_letter(os_idx)}{i}"
        formula_hostname = f'=IF(ISBLANK({h_orig_cell}), IF(ISBLANK(INDEX(Hosts!$A$2:$A${host_rows}, MATCH({ip_orig_cell}, Hosts!$B$2:$B${host_rows}, 0))), "", INDEX(Hosts!$A$2:$A${host_rows}, MATCH({ip_orig_cell}, Hosts!$B$2:$B${host_rows}, 0))), {h_orig_cell})'
        sheet.cell(row=i, column=h_idx).value = formula_hostname
        formula_ip = f'=IF(ISBLANK({ip_orig_cell}), IF(ISBLANK(INDEX(Hosts!$B$2:$B${host_rows}, MATCH({h_orig_cell}, Hosts!$A$2:$A${host_rows}, 0))), "", INDEX(Hosts!$B$2:$B${host_rows}, MATCH({h_orig_cell}, Hosts!$A$2:$A${host_rows}, 0))), {ip_orig_cell})'
        sheet.cell(row=i, column=ip_idx).value = formula_ip
        formula_os = f'=IF(ISBLANK({os_orig_cell}), IF(ISBLANK(INDEX(Hosts!$C$2:$C${host_rows}, MATCH({ip_cell}, Hosts!$B$2:$B${host_rows}, 0))), "", INDEX(Hosts!$C$2:$C${host_rows}, MATCH({ip_cell}, Hosts!$B$2:$B${host_rows}, 0))), {os_orig_cell})'
        sheet.cell(row=i, column=os_idx).value = formula_os


def add_status_validation(sheet):
    """
    Add data validation for the "Status" column on the given sheet.
    Acceptable values: "-", "Open", "On-going", "Closed".
    """
    headers = [cell.value for cell in sheet[1]]
    try:
        status_idx = headers.index("Status") + 1
    except ValueError:
        return
    max_row = sheet.max_row
    dv = DataValidation(type="list", formula1='"-,Open,On-going,Closed"', allow_blank=False)
    dv.error = 'Select a value from the list'
    dv.errorTitle = 'Invalid Entry'
    dv_range = f"{get_column_letter(status_idx)}2:{get_column_letter(status_idx)}{max_row}"
    sheet.add_data_validation(dv)
    dv.add(dv_range)


def add_conditional_formatting(sheet):
    """
    Add conditional formatting to the "Status" column.
    Applies a red fill for "Open" and an orange fill for "On-going".
    """
    headers = [cell.value for cell in sheet[1]]
    try:
        status_idx = headers.index("Status") + 1
    except ValueError:
        return
    max_row = sheet.max_row
    data_range = f"A2:{get_column_letter(sheet.max_column)}{max_row}"
    formula_open = f'=${get_column_letter(status_idx)}2="Open"'
    formula_ongoing = f'=${get_column_letter(status_idx)}2="On-going"'
    light_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    light_orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    sheet.conditional_formatting.add(data_range,
                                     FormulaRule(formula=[formula_open], fill=light_red_fill))
    sheet.conditional_formatting.add(data_range,
                                     FormulaRule(formula=[formula_ongoing], fill=light_orange_fill))


def extract_linux_users(plugin_output):
    """
    Extract individual Linux user names from plugin output.
    """
    linux_re = re.compile(r"^User\s*:\s*(.+)$")
    users = []
    for line in plugin_output.splitlines():
        line = line.strip()
        m = linux_re.match(line)
        if m:
            user = m.group(1).strip()
            if user:
                users.append(user)
    return users


def extract_wmi_users(plugin_output):
    """
    Extract individual WMI user names from plugin output.
    """
    wmi_re = re.compile(r"^Name\s*:\s*(.+)$")
    users = []
    for line in plugin_output.splitlines():
        line = line.strip()
        if line.lower().startswith("no. of users"):
            continue
        m = wmi_re.match(line)
        if m:
            user = m.group(1).strip()
            if user:
                users.append(user)
    return users


def extract_linux_software(plugin_output):
    """
    Extract Linux installed programs from plugin output.
    """
    programs = []
    for line in plugin_output.splitlines():
        line = line.strip()
        if not line:
            continue
        if "list of packages" in line.lower() or line.startswith("-----"):
            continue
        parts = line.split("|")
        if parts:
            prog = parts[0].strip()
            if prog:
                programs.append(prog)
    return programs


def extract_windows_software(plugin_output):
    """
    Extract Windows installed software entries from plugin output.
    """
    programs = []
    for line in plugin_output.splitlines():
        line = line.strip()
        if not line:
            continue
        if "installed on" in line.lower():
            programs.append(line)
    return programs


def extract_users(row):
    """
    Determine and extract users from a row based on its "Name" field.
    """
    plugin_output = row.get("Plugin Output", "")
    name_val = row.get("Name", "").strip()
    if name_val == "Linux User List Enumeration":
        return extract_linux_users(plugin_output)
    elif name_val == "Enumerate Users via WMI":
        return extract_wmi_users(plugin_output)
    else:
        return []


def extract_installed_software(row):
    """
    Determine and extract installed software from a row based on its "Name" field.
    """
    plugin_output = row.get("Plugin Output", "")
    name_val = row.get("Name", "").strip()
    if "ssh" in name_val.lower():
        return extract_linux_software(plugin_output)
    elif "microsoft windows installed software enumeration" in name_val.lower():
        return extract_windows_software(plugin_output)
    else:
        return []


def hide_and_autowidth_columns(sheet, allowed_columns, auto_columns):
    """
    Hide all columns not listed in allowed_columns.
    For columns in auto_columns, adjust the width based on the maximum cell length.
    """
    header = [cell.value for cell in sheet[1]]
    for idx, col_name in enumerate(header, start=1):
        col_letter = get_column_letter(idx)
        if col_name not in allowed_columns:
            sheet.column_dimensions[col_letter].hidden = True
        elif col_name in auto_columns:
            max_length = 0
            for cell in sheet[col_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max_length + 2


def add_risk_font_formatting(sheet, risk_mapping):
    """
    Add conditional formatting rules to change the font color based on the risk level.
    """
    headers = [cell.value for cell in sheet[1]]
    try:
        risk_idx = headers.index("Risk") + 1
    except ValueError:
        return
    risk_col_letter = get_column_letter(risk_idx)
    max_row = sheet.max_row
    data_range = f"{risk_col_letter}2:{risk_col_letter}{max_row}"
    for risk_val, color in risk_mapping.items():
        rule = FormulaRule(formula=[f'=${risk_col_letter}2="{risk_val}"'], font=Font(color=color))
        sheet.conditional_formatting.add(data_range, rule)


def csv_to_excel(csv_filename: str, excel_filename: str, logger=None, software_exclusion_keywords=None):
    """
    Convert the given CSV file into an Excel workbook.
    Six sheets are created (in order):
      1. "Hosts": Unique hosts from the CSV with columns "Hostname", "IP", "OS".
      2. "Vulnerabilities": Processed CSV rows with Risk in {"None","Low","Medium","High","Critical"}.
         After removing "Host" and "Plugin Output", immediately after "Risk" insert six new host-lookup columns:
         "Hostname_original", "Hostname", "IP_original", "IP", "OS_original", "OS", then append "Status" (set to "Open")
         and "Remarks".
      3. "Compliance": Processed similarly for rows with Risk in {"FAILED","WARNING"}.
      4. "Open Ports": Processed similarly for rows where Name equals "Netstat Portscanner (SSH)".
      5. "Users": For rows where Name is "Linux User List Enumeration" or "Enumerate Users via WMI",
         extract individual users from "Plugin Output" and for each create a new row. The header is built as above,
         then extra columns "User", "Status", "Remarks" are appended. Additionally, the six host-lookup columns
         are included and filled with structured formulas.
      6. "Installed Software": For rows where Name contains "Software Enumeration" (case insensitive)
         and does not contain any exclusion keyword (default: ["startup", "start-up", "os"]),
         for each host (only the first matching row per host is kept) the "Plugin Output" is split (by newlines)
         and each extracted line is written as a new row. The header is built as above but with extra columns
         "Installed Programs", "Status", "Remarks". Again, the six host-lookup columns are included and filled.
    
    All CSV-derived sheets are converted into Excel tables with header cells styled with a solid fill (RGB A5A5A5)
    with black text and full thin borders. Data validation is added so that the "Status" column only accepts
    "-", "Open", "On-going", or "Closed". In Vulnerabilities, Compliance and Open Ports the host-lookup columns
    are filled with lookup formulas (using A1-style references to the Hosts sheet). For Users and Installed Software,
    the host-lookup columns are present and then overwritten with structured reference formulas.
    
    Finally, the visible columns are limited (the rest are hidden) as follows:
         - Hosts: ["Hostname", "IP", "OS"]
         - Vulnerabilities: ["Risk", "Hostname", "IP", "OS", "Name", "Synopsis", "Description", "Solution", "See Also", "Status", "Remarks"]
         - Compliance: ["Risk", "Hostname", "IP", "OS", "Name", "Description", "Solution", "See Also", "Status", "Remarks"]
         - Open Ports: ["Hostname", "IP", "OS", "Protocol", "Port", "Status", "Remarks"]
         - Users: ["Hostname", "IP", "OS", "User", "Status", "Remarks"]
         - Installed Software: ["Hostname", "IP", "OS", "Installed Programs", "Status", "Remarks"]
    """
    if software_exclusion_keywords is None:
        software_exclusion_keywords = ["startup", "start-up", "os"]

    # --- Step 1. Read CSV and split into datasets ---
    vulnerabilities = []
    compliance = []
    open_ports = []
    user_enum = []       # For user enumeration rows.
    software_enum_dict = {}  # Group software enumeration rows by host (only first per host)
    hosts_set = set()

    try:
        with open(csv_filename, newline="", encoding="utf-8") as csvfile:
            reader = csv.DictReader(csvfile)
            csv_header = reader.fieldnames  # Includes "Host" and "Plugin Output"
            rows = list(reader)
            rows = remove_duplicates(rows, csv_header)
            for row in rows:
                host_val = row.get("Host", "").strip()
                if host_val:
                    hosts_set.add(host_val)
                risk = row.get("Risk", "").strip()
                name_val = row.get("Name", "").strip()
                if risk in {"None", "Low", "Medium", "High", "Critical"}:
                    vulnerabilities.append(row)
                elif risk in {"FAILED", "WARNING"}:
                    compliance.append(row)
                if name_val == "Netstat Portscanner (SSH)":
                    open_ports.append(row)
                if name_val in {"Linux User List Enumeration", "Enumerate Users via WMI"}:
                    user_enum.append(row)
                if "software enumeration" in name_val.lower():
                    exclude = any(kw.lower() in name_val.lower() for kw in software_exclusion_keywords)
                    if not exclude and host_val and host_val not in software_enum_dict:
                        software_enum_dict[host_val] = row
        software_enum = list(software_enum_dict.values())
    except Exception as e:
        if logger:
            logger.error("Error reading CSV file %s: %s", csv_filename, e)
        return

    # --- Step 2. Create workbook and Hosts sheet ---
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    hosts_sheet = wb.create_sheet("Hosts")
    hosts_header = ["Hostname", "IP", "OS"]
    hosts_sheet.append(hosts_header)
    for host in sorted(hosts_set):
        if IP_REGEX.match(host):
            hosts_sheet.append(["", host, ""])
        else:
            hosts_sheet.append([host, "", ""])

    create_table_and_style(hosts_sheet, "Hosts")

    # --- Step 3. Create CSV-derived sheets ---
    vuln_sheet = wb.create_sheet("Vulnerabilities")
    write_csv_sheet(vuln_sheet, vulnerabilities, csv_header)
    comp_sheet = wb.create_sheet("Compliance")
    write_csv_sheet(comp_sheet, compliance, csv_header)
    open_ports_sheet = wb.create_sheet("Open Ports")
    write_csv_sheet(open_ports_sheet, open_ports, csv_header)
    users_sheet = wb.create_sheet("Users")
    write_csv_sheet(users_sheet, user_enum, csv_header, "Users", extract_users)
    installed_software_sheet = wb.create_sheet("Installed Software")
    write_csv_sheet(installed_software_sheet, software_enum, csv_header, "Installed Programs", extract_installed_software)

    create_table_and_style(vuln_sheet, "Vulnerabilities")
    create_table_and_style(comp_sheet, "Compliance")
    create_table_and_style(open_ports_sheet, "Open Ports")
    create_table_and_style(users_sheet, "Users")
    create_table_and_style(installed_software_sheet, "Installed Software")

    # --- Step 4. Add lookup formulas, validation, and conditional formatting ---
    add_lookup_formulas(vuln_sheet, hosts_sheet)
    add_lookup_formulas(comp_sheet, hosts_sheet)
    add_lookup_formulas(open_ports_sheet, hosts_sheet)
    add_lookup_formulas(users_sheet, hosts_sheet)
    add_lookup_formulas(installed_software_sheet, hosts_sheet)

    add_status_validation(vuln_sheet)
    add_status_validation(comp_sheet)
    add_status_validation(open_ports_sheet)
    add_status_validation(users_sheet)
    add_status_validation(installed_software_sheet)

    add_conditional_formatting(vuln_sheet)
    add_conditional_formatting(comp_sheet)
    add_conditional_formatting(open_ports_sheet)
    add_conditional_formatting(users_sheet)
    add_conditional_formatting(installed_software_sheet)

    # --- Step 5. Hide unwanted columns ---
    hide_and_autowidth_columns(hosts_sheet, ["Hostname", "IP", "OS"], [])
    hide_and_autowidth_columns(vuln_sheet, ["Risk", "Hostname", "IP", "OS", "Name", "Synopsis", "Description", "Solution", "See Also", "Status", "Remarks"], [])
    hide_and_autowidth_columns(comp_sheet, ["Risk", "Hostname", "IP", "OS", "Name", "Description", "Solution", "See Also", "Status", "Remarks"], [])
    hide_and_autowidth_columns(open_ports_sheet, ["Hostname", "IP", "OS", "Protocol", "Port", "Status", "Remarks"], [])
    hide_and_autowidth_columns(users_sheet, ["Hostname", "IP", "OS", "User", "Status", "Remarks"], [])
    hide_and_autowidth_columns(installed_software_sheet, ["Hostname", "IP", "OS", "Installed Programs", "Status", "Remarks"], [])

    # --- Step 6. Add risk font conditional formatting ---
    vuln_risk_mapping = {"Critical": "8B0000", "High": "FF0000", "Medium": "A0522D", "Low": "0000FF"}
    add_risk_font_formatting(vuln_sheet, vuln_risk_mapping)
    comp_risk_mapping = {"FAILED": "FF0000", "WARNING": "A0522D"}
    add_risk_font_formatting(comp_sheet, comp_risk_mapping)

    # --- Step 7. Save the workbook ---
    try:
        wb.save(excel_filename)
        if logger:
            logger.info("Excel file created successfully at: %s", excel_filename)
    except Exception as e:
        if logger:
            logger.error("Error saving Excel file %s: %s", excel_filename, e)